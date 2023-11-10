import asyncio
from flask import Flask, render_template,request, jsonify, send_file
from flask_socketio import SocketIO, emit
from openpyxl import Workbook
app = Flask(__name__)
socketio = SocketIO(app)

import sqlite3
DATABASE = 'db.sqlite3'


def create_tables():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS node_data (
            id INTEGER PRIMARY KEY,
            name TEXT,
            left TEXT,
            top TEXT,
            states TEXT,
            children TEXT,
            "values" TEXT,
            parents TEXT
        )
    ''')

    # Insert nodeData into the node_data table
    for node in nodeData:
        cursor.execute('''
            INSERT INTO node_data (name, left, top, states, children, "values", parents)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            node.get('name', ''),
            node.get('left', ''),
            node.get('top', ''),
            json.dumps(node.get('states', [])),
            json.dumps(node.get('children', [])),
            json.dumps(node.get('values', [])),
            json.dumps(node.get('parents', [])),
        ))

    conn.commit()
    conn.close()

nodeData = [
    {
        "name": "Fish",
        "left": "1246px",
        "top": "353px",
        "states": [
            {
                "name": "zero",
                "probability": 0.2
            },
            {
                "name": "low",
                "probability": 0
            },
            {
                "name": "medium",
                "probability": 0
            },
            {
                "name": "high",
                "probability": 0
            },
            {
                "name": "zero",
                "probability": 0.2
            }
        ],
        "children": []
    },
    {
        "name": "Cover",
        "left": "717px",
        "top": "444px",
        "states": [
            {
                "name": "zero",
                "probability": 1
            },
            {
                "name": "low",
                "probability": 0.25
            },
            {
                "name": "medium",
                "probability": 0.01
            },
            {
                "name": "high",
                "probability": 0.5
            }
        ],
        "children": [
            "Fish"
        ]
    },
    {
        "name": "Migration",
        "left": "689px",
        "top": "107px",
        "states": [
            {
                "name": "zero",
                "probability": 0.95
            },
            {
                "name": "low",
                "probability": 0.5
            },
            {
                "name": "medium",
                "probability": 0.25
            },
            {
                "name": "high",
                "probability": 0
            }
        ],
        "children": [
            "Fish"
        ]
    },
    {
        "name": "Barrier",
        "left": "62px",
        "top": "112px",
        "states": [
            {
                "name": "none",
                "probability": 0.3
            },
            {
                "name": "some",
                "probability": 0.5
            },
            {
                "name": "many",
                "probability": 0.2
            },
            {
                "name": "huge",
                "probability": 0.2
            }
        ],
        "children": [
            "Migration"
        ]
    },
    {
        "name": "Discharge",
        "left": "72px",
        "top": "441px",
        "states": [
            {
                "name": "zero",
                "probability": 0.1
            },
            {
                "name": "low",
                "probability": 0.4
            },
            {
                "name": "medium",
                "probability": 0.4
            },
            {
                "name": "high",
                "probability": 0.1
            }
        ],
        "children": [
            "Cover"
        ]
    }
]
import json
parent_child_map = {}

# Populate parent_child_map with parent-child relationships
for node in nodeData:
    if "children" in node and node["children"]:
        for child_name in node["children"]:
            if child_name in parent_child_map:
                parent_child_map[child_name].append(node["name"])
            else:
                parent_child_map[child_name] = [node["name"]]

# Update nodes with values and parents
for node in nodeData:
    node["values"] = [[], []]
    if node["name"] in parent_child_map:
        node["parents"] = parent_child_map[node["name"]]
    else:
        node["parents"] = []

# Print the formatted JSON using json.dumps()
formatted_json = json.dumps(nodeData, indent=4)
print(formatted_json)

# sqlite database connection
create_tables()

@app.route("/")
def hello_world():
    return render_template("main.html")

@app.route("/water")
def water():
    return render_template("water.html")

@socketio.on('connect')
def on_connect():
    print('A client connected')

@socketio.on('disconnect')
def on_disconnect():
    print('A client disconnected')

@socketio.on('client_message')
def handle_client_message(data):
    print('Received from client:', data['message'])
    # Here you can process the data or send a response back to the client
    # For example:
    response_message = 'Message received on the server!'
    emit('server_message', {'message': response_message})


@socketio.on('request_node_data')
def send_node_data():
    emit('node_data', {'data': nodeData})


@socketio.on('update_node_data')
def handle_update_node_data(data):
    updated_node_data = data['data']
    global nodeData
    nodeData = updated_node_data
    # Emit the updated data to all connected clients
    emit('node_data', {'data': nodeData}, broadcast=True)
    
import json

@app.route('/api/generate_excel', methods=['POST'])
def generate_excel():
    try:
        data = request.get_json()
        string1 = data.get('string1', '')
        string2 = data.get('string2', '')

        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        nodes = string1.split(",")
        values = string2.replace("\"","").replace(":"," ").split("||")
        # Place the strings in the first and second rows
        nodes += "DISCHARGE_YR++DISCHARGE_LF++DISCHARGE_HF++DISCHARGE_FD".split("++")
        values += "{0 0.05, 0.9 0.2, 4.6 0.5, 20.9 0.2, 500.6 0.05}++{0 0.05, 0 0.2, 0 0.5, 0.9 0.2, 1.8 0.05}++{1.8 0.05, 4.6 0.2, 11.1 0.5, 20.9 0.2, 11.1 0.05}++{156.3 0.05, 233.1 0.2, 277.9 0.5, 438.2 0.2, 971.6 0.05}".split("++")
        
        for i in range(len(nodes)):
          sheet.cell(row=1, column=i+1, value=nodes[i])
          sheet.cell(row=2, column=i+1, value=str(values[i]))

        # Save the workbook to a temporary file
        excel_file = 'netica_case.xlsx'
        workbook.save(excel_file)

        # Send the Excel file for download
        return send_file(excel_file, as_attachment=True, download_name='netica_case.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': 'An error occurred while generating the Excel file'}), 500
    
if __name__ == "__main__":
       
    # socketio.run(app, debug=True, port=4400,allow_unsafe_werkzeug=True)
    import os
    from hypercorn.config import Config
    from hypercorn.asyncio import serve

    port = int(os.environ.get('PORT', 4400))  # Default to 4400 if PORT environment variable not set
    config = Config()
    config.bind = [f"0.0.0.0:{port}"]
    
    
    asyncio.run(serve(app, config))

   